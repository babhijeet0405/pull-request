import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from copy import copy
import os
import platform
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# Print both current working directory and script directory for comparison
print('Current working directory:', os.getcwd())

# Get the directory where this script is located
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
print('Script directory:', BASE_DIR)

# Use the script directory to ensure the Excel template is found
EXCEL_TEMPLATE = os.path.join(BASE_DIR, "Standard_Template_new.xlsx")
print('Excel template path:', EXCEL_TEMPLATE)

# Injection pressure mapping
injection_pressure_map = {
    "ABS UV": 700, "ABS Non UV": 700, "PP T20": 650, "PC Clear": 900,
    "POM": 800, "PC Amber": 850, "PC black": 850, "PMMA Clear": 780,
    "PMMA Red": 780, "PP+EPDM": 720, "PA66": 950, "EPDM": 750, "PP": 650,
    "PC ABS": 820, "PP EPDM": 700, "PP HD": 670, "PA6": 940,
    "PA66 GF30": 1000, "PA66 GF15": 980, "PA66 GF25": 990,
    "PP GF20": 950, "PP GF10": 900, "PA6 GF30": 990, "PA6 GF15": 970,
    "PA6 GF25": 980, "PP/PE-T16": 720, "PP EPDM TD10": 710,
    "SEBS": 680, "HDPE": 650, "PE": 640
}

# Dropdowns
dropdowns = {
    "In-house/Purchase": ["In-House", "Purchase"],
    "Rawmaterial": list(injection_pressure_map.keys()),
    "No. of cavities -X direction": ["1", "2", "4", "6", "8", "12", "16"],
    "No. of cavities -Y direction": ["1", "2", "4", "6", "8", "12", "16"],
    "Grain requirement": ["Y", "N"],
    "HRS Type": ["0", "1", "2", "3"],
    "No.of drop": [str(i) for i in range(1, 13)],
    "X Direction Slider": ["Y", "N"],
    "Y Direction Slider": ["Y", "N"],
    "No.of Sliders/Cavity": [str(i) for i in range(12)],
    "No.of Lifters/cavity": [str(i) for i in range(12)],
    "No.of Bosses /cavity": [str(i) for i in range(12)],
    "No.of Ejector pads/ cavity": [str(i) for i in range(12)],
    "Projected Area %": ["50%", "55%", "60%", "70%", "75%", "80%", "85%", "90%", "100%"],
    "Process": ["Injection moulding", "Blow moulding", "Assy", "BOP"],
    "M/c selected": [str(i) for i in [60, 90, 120, 160, 200, 220, 250, 260, 280, 320, 380,
                                      470, 530, 600, 700, 800, 900, 1000, 1200, 1300, 1400,
                                      1600, 1850, 2100, 2400, 2800, 3300, 4000]],
    "Labor Type (Machine)": ["Skilled", "Semi skilled", "Unskilled"],
    "Labor Type (Setup)": ["Skilled", "Semi skilled", "Unskilled"]
}

fields = [
    "In-house/Purchase", "ZF Part Number", "Part No", "Part Name", "Volume P.A Nos",
    "No.off (per Vehicle)", "Rawmaterial", "L(mm)", "W(mm)", "H(mm)", "T(mm)",
    "No. of cavities -X direction", "No. of cavities -Y direction", "Grain requirement",
    "HRS Type", "No.of drop", "X Direction Slider", "Y Direction Slider",
    "No.of Sliders/Cavity", "No.of Lifters/cavity", "No.of Bosses /cavity",
    "No.of Ejector pads/ cavity", "Projected Area %", "Injection Pressure", "Total Projected Area", "Finish wt.( Kgs)",
    "Scrap cost/Kg (INR)", "Bop Parts", "Qty (Nos)", "Cost /part (INR)", "Process", "M/c selected", "Machine Cycle time (sec)",
    "Labor Type (Machine)", "Operator /machine", "Labor Type (Setup)", "No.of operator", "Time/Lot (min)"
]

all_fields = ["Sl.#"] + [f for f in fields if f != "Injection Pressure"]

numeric_fields = {
    "L(mm)": float, "W(mm)": float, "H(mm)": float, "T(mm)": float,
    "Volume P.A Nos": int, "No.off (per Vehicle)": int,
    "No. of cavities -X direction": int, "No. of cavities -Y direction": int, "HRS Type": int,
    "No.of drop": int, "No.of Sliders/Cavity": int, "No.of Lifters/cavity": int,
    "No.of Bosses /cavity": int, "No.of Ejector pads/ cavity": int,
    "Finish wt.( Kgs)": float, "Scrap cost/Kg (INR)": float, "Bop Parts": float,
    "Qty (Nos)": int, "Cost /part (INR)": float, "M/c selected": int, "Machine Cycle time (sec)": float,
    "Operator /machine": int, "No.of operator": int, "Time/Lot (min)": float
}

section_titles = {
    "Raw material cost": ["Finish wt.( Kgs)", "Scrap cost/Kg (INR)"],
    "BOP parts cost": ["Bop Parts", "Qty (Nos)", "Cost /part (INR)"],
    "Machine cost": ["Process", "M/c selected", "Machine Cycle time (sec)"],
    "Labor Cost": ["Labor Type (Machine)", "Operator /machine"],
    "Setting-up cost": ["Labor Type (Setup)", "No.of operator", "Time/Lot (min)"],
}

export_mode = None

def get_section(field):
    for section, items in section_titles.items():
        if field in items:
            return section
    return None

def update_tooling_by_zf_part_number(wb, zf_part):
    try:
        entries_ws = wb["Entries"]
        tooling_ws = wb["Tooling"]
        calculation_ws = wb["Calculation"]

        found_entries_row = None
        found_calc_row = None
        
        # Find the correct row in Entries sheet
        for row in range(3, entries_ws.max_row + 1):
            cell_val = entries_ws.cell(row=row, column=3).value  # Column C = ZF Part Number
            if cell_val and str(cell_val).strip().lower() == zf_part.strip().lower():
                found_entries_row = row
                print(f"Tooling update: Found ZF Part '{zf_part}' in Entries at row {found_entries_row}")
                break

        if not found_entries_row:
            messagebox.showerror("Not Found", f"ZF Part Number '{zf_part}' not found in Entries sheet.")
            return False

        # Find the correct row in Calculation sheet
        for row in range(6, calculation_ws.max_row + 1):
            cell_val = calculation_ws.cell(row=row, column=3).value  # Column C = ZF Part Number
            if cell_val and str(cell_val).strip().lower() == zf_part.strip().lower():
                found_calc_row = row
                print(f"Tooling update: Found ZF Part '{zf_part}' in Calculation at row {found_calc_row}")
                break

        # Copy data from found entries row to tooling row 6 (columns A to AA)
        for col in range(1, 28):  # A to AA
            tooling_ws.cell(row=6, column=col).value = entries_ws.cell(row=found_entries_row, column=col).value
        
        print(f"Tooling update: Copied data from Entries row {found_entries_row} to Tooling row 6")
        
        # Verify the part name was copied correctly
        part_name_tooling = tooling_ws.cell(row=6, column=4).value  # Column D = Part Name
        print(f"Tooling update: Part name in tooling sheet is now: {part_name_tooling}")

        # Copy calculation data if found (AP (42) → AB (28), AS (45) → AC (29))
        if found_calc_row:
            tooling_ws.cell(row=6, column=28).value = calculation_ws.cell(row=found_calc_row, column=42).value
            tooling_ws.cell(row=6, column=29).value = calculation_ws.cell(row=found_calc_row, column=45).value

        return True
    except Exception as e:
        messagebox.showerror("Error", f"Failed to update tooling sheet.\n{e}")
        return False

def export_with_formulas(zf_part):
    try:
        print(f">>> Starting export WITH formulas for: {zf_part}")
        wb = load_workbook(EXCEL_TEMPLATE, data_only=False)
        if export_logic(wb, zf_part, export_type="with"):
            print("✅ Export with formulas completed successfully")
    except Exception as e:
        messagebox.showerror("Export Error", str(e))
        print("❌ Export with formulas failed:", e)

def export_without_formulas(zf_part):
    try:
        print(f">>> Starting export WITHOUT formulas for: {zf_part}")
        wb = load_workbook(EXCEL_TEMPLATE, data_only=True)
        if export_logic(wb, zf_part, export_type="without"):
            print("✅ Export without formulas completed successfully")
    except Exception as e:
        messagebox.showerror("Export Error", str(e))
        print("❌ Export without formulas failed:", e)


def export_logic(wb, zf_part, export_type="with"):
    export_wb = Workbook()
    export_wb.remove(export_wb.active)

    # Debug: Print all available sheet names
    print(f"Available sheets in workbook: {wb.sheetnames}")

    calc_ws = wb["Calculation"]
    tooling_ws = wb["Tooling"]
    summary_ws = wb["Summary"] if "Summary" in wb.sheetnames else None

    found_row = None
    found_entries_row = None

    # 🟡 STEP 1: Find row by ZF Part Number in Calculation sheet
    for row in range(6, calc_ws.max_row + 1):
        val = calc_ws[f"C{row}"].value
        if val and str(val).strip().lower() == zf_part.strip().lower():
            found_row = row
            print(f"Found ZF Part '{zf_part}' in Calculation sheet at row {found_row}")
            break

    if not found_row and export_type == "with":
        try:
            temp_wb = load_workbook(EXCEL_TEMPLATE, data_only=True)
            temp_calc = temp_wb["Calculation"]
            for row in range(6, temp_calc.max_row + 1):
                val = temp_calc[f"C{row}"].value
                if val and str(val).strip().lower() == zf_part.strip().lower():
                    found_row = row
                    print(f"Fallback: Found ZF Part '{zf_part}' in Calculation sheet at row {found_row}")
                    break
        except Exception as e:
            messagebox.showerror("Lookup Error", f"Fallback lookup failed:\n{str(e)}")
            return False

    if not found_row:
        messagebox.showerror("Not Found", f"ZF Part Number '{zf_part}' not found in Calculation sheet.")
        return False

    # 🟡 STEP 2: Find corresponding row in Entries sheet
    entries_ws = wb["Entries"] if "Entries" in wb.sheetnames else None
    if entries_ws:
        for row in range(3, entries_ws.max_row + 1):
            val = entries_ws[f"C{row}"].value
            if val and str(val).strip().lower() == zf_part.strip().lower():
                found_entries_row = row
                print(f"Found ZF Part '{zf_part}' in Entries sheet at row {found_entries_row}")
                break

    # 🟡 STEP 3: Update tooling sheet with correct data BEFORE export
    if found_entries_row and not update_tooling_by_zf_part_number(wb, zf_part):
        return False

    # 🟡 STEP 3.5: Update Summary sheet with the same data as Tooling
    if found_entries_row and summary_ws:
        try:
            print(f"Updating Summary sheet with data from Entries row {found_entries_row}")
            
            # Instead of copying from Tooling, let's copy directly from Entries with proper column mapping for Summary
            entries_ws_ref = wb["Entries"]
            
            # Define the correct column mapping for Summary sheet
            # Map from Entries columns to Summary columns based on actual column headers
            summary_column_mapping = {
                # Entries -> Summary mapping
                1: 1,   # Sl.# (Entries A) -> Sl.# (Summary A)
                3: 2,   # ZF Part Number (Entries C) -> ZF Part Number (Summary B)
                5: 3,   # Part Name (Entries E) -> Part Name (Summary C)
                7: 4,   # No.off (per Vehicle) (Entries G) -> No.off (per Vehicle) (Summary D)
                6: 5,   # Volume P.A Nos (Entries F) -> Volume P.A Nos (Summary E)
                8: 6,   # Rawmaterial (Entries H) -> Rawmaterial (Summary F)
                # Part Size not in Entries, will be calculated/set separately
                10: 8,  # W(mm) (Entries J) -> W(mm) (Summary H)
                11: 9,  # H(mm) (Entries K) -> H(mm) (Summary I)
                12: 10, # T(mm) (Entries L) -> T(mm) (Summary J)
                # Number of Cavity = X direction * Y direction cavities from Entries
                # Will be calculated separately
                # Other fields like costs will be calculated by the application logic
            }
            
            # Special calculation fields that need to be computed
            # Number of Cavity = X direction * Y direction
            x_cavities = entries_ws_ref.cell(row=found_entries_row, column=13).value or 1  # No. of cavities -X direction
            y_cavities = entries_ws_ref.cell(row=found_entries_row, column=14).value or 1  # No. of cavities -Y direction
            total_cavities = int(x_cavities) * int(y_cavities)
            summary_ws.cell(row=6, column=11).value = total_cavities  # Number of Cavity Nos
            
            # Part Size calculation (L x W x H from Entries)
            l_mm = entries_ws_ref.cell(row=found_entries_row, column=9).value or 0   # L(mm)
            w_mm = entries_ws_ref.cell(row=found_entries_row, column=10).value or 0  # W(mm) 
            h_mm = entries_ws_ref.cell(row=found_entries_row, column=11).value or 0  # H(mm)
            if l_mm and w_mm and h_mm:
                part_size = f"{l_mm} x {w_mm} x {h_mm}"
                summary_ws.cell(row=6, column=7).value = part_size  # Part Size (As per tooling direction)
            
            # Copy data using the correct column mapping
            for entries_col, summary_col in summary_column_mapping.items():
                summary_ws.cell(row=6, column=summary_col).value = entries_ws_ref.cell(row=found_entries_row, column=entries_col).value
            
            print("Summary sheet updated with data from Entries using correct column mapping")
            
            # Verify the data was copied correctly - checking actual Summary sheet columns
            sl_no_summary = summary_ws.cell(row=6, column=1).value      # Column A = Sl.#
            zf_part_summary = summary_ws.cell(row=6, column=2).value    # Column B = ZF Part Number  
            part_name_summary = summary_ws.cell(row=6, column=3).value  # Column C = Part Name
            no_off_summary = summary_ws.cell(row=6, column=4).value     # Column D = No.off (per Vehicle)
            volume_summary = summary_ws.cell(row=6, column=5).value     # Column E = Volume P.A Nos
            rawmaterial_summary = summary_ws.cell(row=6, column=6).value # Column F = Rawmaterial
            part_size_summary = summary_ws.cell(row=6, column=7).value  # Column G = Part Size
            w_mm_summary = summary_ws.cell(row=6, column=8).value       # Column H = W(mm)
            h_mm_summary = summary_ws.cell(row=6, column=9).value       # Column I = H(mm)
            t_mm_summary = summary_ws.cell(row=6, column=10).value      # Column J = T(mm)
            cavities_summary = summary_ws.cell(row=6, column=11).value  # Column K = Number of Cavity Nos
            
            print(f"Summary sheet verification:")
            print(f"  Sl.#: {sl_no_summary}")
            print(f"  ZF Part Number: {zf_part_summary}")
            print(f"  Part Name: {part_name_summary}")
            print(f"  No.off (per Vehicle): {no_off_summary}")
            print(f"  Volume P.A Nos: {volume_summary}")
            print(f"  Rawmaterial: {rawmaterial_summary}")
            print(f"  Part Size: {part_size_summary}")
            print(f"  W(mm): {w_mm_summary}")
            print(f"  H(mm): {h_mm_summary}")
            print(f"  T(mm): {t_mm_summary}")
            print(f"  Number of Cavity Nos: {cavities_summary}")
            
        except Exception as e:
            print(f"Warning: Failed to update Summary sheet: {e}")

    # Refresh the tooling worksheet reference after update
    tooling_ws = wb["Tooling"]
    
    # Get the updated tooling values and copy to calculation
    calc_ws[f"AQ{found_row}"] = tooling_ws["AD6"].value
    calc_ws[f"AR{found_row}"] = tooling_ws["AE6"].value

    # Find corresponding summary row (after we updated it to row 6)
    found_summary_row = 6  # We always update Summary at row 6, same as Tooling
    if summary_ws:
        print(f"Using Summary sheet row 6 (updated with current part data)")
    else:
        print("Warning: Summary sheet not found, using calculation row number")
        found_summary_row = found_row

    sheet_rows = {"Calculation": found_row, "Tooling": 6, "Summary": found_summary_row}
    sheet_headers = {"Calculation": [3, 4, 5], "Tooling": [3, 4, 5], "Summary": [3, 4, 5]}
    sheet_col_limits = {
        "Calculation": 90,
        "Tooling": 40,
        "Summary": 39,
        "Import": 28,
        "Master Data": 18,
        "Entries": 38
    }
    
    print(f"Export data rows: Calculation={found_row}, Tooling=6, Summary={found_summary_row}")
    
    # Debug: Show what's in Summary row 6 after our update
    if summary_ws and found_summary_row == 6:
        sl_no_summary = summary_ws.cell(row=6, column=1).value      # Column A = Sl.#
        zf_part_summary = summary_ws.cell(row=6, column=2).value    # Column B = ZF Part Number  
        part_name_summary = summary_ws.cell(row=6, column=3).value  # Column C = Part Name
        no_off_summary = summary_ws.cell(row=6, column=4).value     # Column D = No.off (per Vehicle)
        print(f"Final Summary verification - Sl.#: {sl_no_summary}, ZF Part: {zf_part_summary}, Part Name: {part_name_summary}, No.off: {no_off_summary}")

    for sheet_name in ["Calculation", "Tooling", "Summary", "Master Data", "Entries", "Import"]:
        # Check if the sheet exists before trying to access it
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in workbook, skipping...")
            continue
            
        source = wb[sheet_name]
        target = export_wb.create_sheet(sheet_name)
        # Hide Master Data and Entries sheets in exported file
        if sheet_name in ["Master Data", "Entries"]:
            target.sheet_state = 'hidden'

        if sheet_name == "Master Data":
            for row in range(1, 111):  # Rows 1 to 110
                for col in range(1, 19):  # Columns A to R
                    cell = source.cell(row=row, column=col)
                    tgt = target.cell(row=row, column=col, value=cell.value)
                    if cell.has_style:
                        tgt.font = copy(cell.font)
                        tgt.border = copy(cell.border)
                        tgt.fill = copy(cell.fill)
                        tgt.number_format = copy(cell.number_format)
                        tgt.protection = copy(cell.protection)
                        tgt.alignment = copy(cell.alignment)
            for col_idx in range(1, 19):
                col_letter = get_column_letter(col_idx)
                if col_letter in source.column_dimensions:
                    target.column_dimensions[col_letter].width = source.column_dimensions[col_letter].width
            for m in source.merged_cells.ranges:
                target.merge_cells(str(m))
            continue

        elif sheet_name == "Entries":
            for row in source.iter_rows(min_row=1, max_row=source.max_row, min_col=1, max_col=38):
                for cell in row:
                    tgt = target.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        tgt.font = copy(cell.font)
                        tgt.border = copy(cell.border)
                        tgt.fill = copy(cell.fill)
                        tgt.number_format = copy(cell.number_format)
                        tgt.protection = copy(cell.protection)
                        tgt.alignment = copy(cell.alignment)
            for col_idx in range(1, 39):
                col_letter = get_column_letter(col_idx)
                if col_letter in source.column_dimensions:
                    target.column_dimensions[col_letter].width = source.column_dimensions[col_letter].width
            for m in source.merged_cells.ranges:
                target.merge_cells(str(m))
            continue

        elif sheet_name == "Import":
            # For Import sheet, we always need to preserve formulas so they can reference the updated Sl.#
            # Load the source with formulas preserved regardless of export_type
            if export_type == "without":
                # For "without formulas" export, we need to temporarily load with formulas preserved
                # to ensure VLOOKUP formulas work with the updated Sl.#, then evaluate them
                temp_wb = load_workbook(EXCEL_TEMPLATE, data_only=False)
                temp_source = temp_wb["Import"]
            else:
                temp_source = source
            
            # Copy all rows and columns to preserve formulas and structure
            for row in temp_source.iter_rows(min_row=1, max_row=temp_source.max_row, min_col=1, max_col=28):
                for cell in row:
                    tgt = target.cell(row=cell.row, column=cell.column)
                    
                    # For cell A2 (Sl.#), set the correct Sl.# for the requested part
                    if cell.row == 2 and cell.column == 1:
                        # Find the Sl.# from the Entries sheet for the requested ZF part
                        if found_entries_row:
                            entries_sl_no = entries_ws.cell(row=found_entries_row, column=1).value  # Sl.# from Entries
                            print(f"Import sheet: Setting Sl.# to {entries_sl_no} for ZF Part '{zf_part}'")
                            tgt.value = entries_sl_no
                        else:
                            tgt.value = cell.value
                    else:
                        # Copy the original value (formulas will reference the updated A2)
                        if export_type == "with":
                            # Keep formulas for "with formulas" export
                            tgt.value = cell.value
                        else:
                            # For "without formulas" export, check if it's a formula
                            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                                # For VLOOKUP formulas, keep the formula so it can be evaluated with the new Sl.#
                                # The formula will be evaluated when the file is saved/opened
                                tgt.value = cell.value
                            else:
                                # For non-formula cells, copy the value directly
                                tgt.value = cell.value
                    
                    # Copy cell styling
                    if cell.has_style:
                        tgt.font = copy(cell.font)
                        tgt.border = copy(cell.border)
                        tgt.fill = copy(cell.fill)
                        tgt.number_format = copy(cell.number_format)
                        tgt.protection = copy(cell.protection)
                        tgt.alignment = copy(cell.alignment)
            
            # Close temporary workbook if we created one
            if export_type == "without":
                temp_wb.close()
            
            # Copy column widths
            for col_idx in range(1, 29):
                col_letter = get_column_letter(col_idx)
                if col_letter in source.column_dimensions:
                    target.column_dimensions[col_letter].width = source.column_dimensions[col_letter].width
            
            # Copy merged cells
            for m in source.merged_cells.ranges:
                target.merge_cells(str(m))
            
            # Debug: Verify Import sheet data after update
            if found_entries_row:
                import_sl_no = target.cell(row=2, column=1).value
                print(f"Import sheet verification:")
                print(f"  Sl.# in A2: {import_sl_no}")
                print(f"  This should match the Sl.# from Entries row {found_entries_row}")
                print(f"  Import sheet formulas will now reference this Sl.# to pull correct data from Calculation sheet")
                
                # Verify some key formulas are preserved
                b2_formula = target.cell(row=2, column=2).value
                f2_formula = target.cell(row=2, column=6).value
                j2_formula = target.cell(row=2, column=10).value
                print(f"  Key formulas preserved:")
                print(f"    B2 (Part Number): {b2_formula}")
                print(f"    F2 (Part Name): {f2_formula}")
                print(f"    J2 (Quantity): {j2_formula}")
            continue

        # ✅ Logic for Calculation, Tooling, Summary (one row + headers)
        max_col = sheet_col_limits[sheet_name]
        data_row = sheet_rows[sheet_name]
        
        print(f"Processing sheet '{sheet_name}': copying data from row {data_row} to export row 6")

        # Copy headers (rows 3, 4, 5)
        for row_idx in sheet_headers[sheet_name]:
            for col in range(1, max_col + 1):
                cell = source.cell(row=row_idx, column=col)
                tgt = target.cell(row=row_idx, column=col, value=cell.value)
                if cell.has_style:
                    tgt.font = copy(cell.font)
                    tgt.border = copy(cell.border)
                    tgt.fill = copy(cell.fill)
                    tgt.number_format = copy(cell.number_format)
                    tgt.protection = copy(cell.protection)
                    tgt.alignment = copy(cell.alignment)

        # Copy the actual data row to row 6 in export
        for col in range(1, max_col + 1):
            src = source.cell(row=data_row, column=col)
            # Get the actual value (whether formula result or raw value)
            if export_type == "with" and src.data_type == 'f':
                # Keep formula for "with formulas" export
                cell_value = src.value
            else:
                # Get calculated value for "without formulas" export or non-formula cells
                cell_value = src.value
            
            tgt = target.cell(row=6, column=col, value=cell_value)
            if src.has_style:
                tgt.font = copy(src.font)
                tgt.border = copy(src.border)
                tgt.fill = copy(src.fill)
                tgt.number_format = copy(src.number_format)
                tgt.protection = copy(src.protection)
                tgt.alignment = copy(src.alignment)

        # Copy merged cells for headers
        for m in source.merged_cells.ranges:
            if (m.min_row in sheet_headers[sheet_name] or m.max_row in sheet_headers[sheet_name]):
                target.merge_cells(str(m))

        # Special handling for Tooling sheet - copy additional rows
        if sheet_name == "Tooling":
            # Copy merged cells for the additional rows
            for m in source.merged_cells.ranges:
                if 8 <= m.min_row <= 48:
                    target.merge_cells(str(m))
            
            # Copy rows 8-48 for tooling sheet
            for r in range(8, 49):
                for c in range(1, max_col + 1):
                    cell = source.cell(row=r, column=c)
                    tgt = target.cell(row=r, column=c, value=cell.value)
                    if cell.has_style:
                        tgt.font = copy(cell.font)
                        tgt.border = copy(cell.border)
                        tgt.fill = copy(cell.fill)
                        tgt.number_format = copy(cell.number_format)
                        tgt.protection = copy(cell.protection)
                        tgt.alignment = copy(cell.alignment)

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            src_width = source.column_dimensions[col_letter].width

            # If original width is None or too small, set a fallback width (like 15)
            target.column_dimensions[col_letter].width = src_width if (src_width and src_width > 6) else 15

    # ✅ Save the exported file
    safe_name = zf_part.replace(" ", "_").replace("/", "_").replace("\\", "_")
    suffix = "WithFormulas" if export_type == "with" else "ValuesOnly"
    export_filename = f"{safe_name}_Export_{suffix}.xlsx"
    
    try:
        export_wb.save(export_filename)
        messagebox.showinfo("Success", f"Exported to '{export_filename}'")
        
        if platform.system() == "Windows":
            os.startfile(export_filename)
        
        return True
    except Exception as e:
        messagebox.showerror("Save Error", f"Failed to save export file: {str(e)}")
        return False

def save_to_standard_excel(entry_data):
    try:
        wb = load_workbook(EXCEL_TEMPLATE) if os.path.exists(EXCEL_TEMPLATE) else Workbook()

        # Create Entries sheet if it doesn't exist
        if "Entries" not in wb.sheetnames:
            ws = wb.create_sheet("Entries")
            for col_index, field in enumerate(all_fields, start=1):
                ws.cell(row=2, column=col_index).value = field
                ws.cell(row=2, column=col_index).font = Font(bold=True)
                ws.cell(row=2, column=col_index).alignment = Alignment(horizontal="center")
        else:
            ws = wb["Entries"]
            ws.data_validations.dataValidation = []

        zf_part = entry_data.get("ZF Part Number", "").strip()
        part_no = entry_data.get("Part No", "").strip()
        overwrite_row = None

        # Look for existing row to edit
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
            zf_value = str(row[2].value).strip() if row[2].value else ""
            part_value = str(row[3].value).strip() if row[3].value else ""
            if zf_value == zf_part and part_value == part_no:
                overwrite_row = row_idx
                break

        # If match found, update that row; else add new row
        if overwrite_row:
            target_row = overwrite_row
            sl_no = target_row - 2
            entry_data["Sl.#"] = sl_no
        else:
            target_row = ws.max_row + 1 if ws.max_row >= 2 else 3
            sl_no = target_row - 2
            entry_data["Sl.#"] = sl_no

        # Write to Entries
        for col_index, field in enumerate(all_fields, start=1):
            ws.cell(row=target_row, column=col_index).value = entry_data.get(field, "")

        # Transfer to Calculation sheet with formulas
        if "Calculation" not in wb.sheetnames:
            wb.create_sheet("Calculation")
        calc_ws = wb["Calculation"]
        calc_ws.data_validations.dataValidation = []

        entries_to_calculation = {
            "A": "A", "B": "B", "C": "C", "D": "D", "E": "E", "F": "F", "G": "G", "H": "H", "I": "I",
            "J": "J", "K": "K", "L": "L", "M": "M", "N": "N", "O": "P", "P": "Q", "Q": "R", "R": "S",
            "S": "T", "T": "U", "U": "V", "V": "W", "W": "X", "X": "Z", "Y": "AA", "Z": "AC", "AA": "AH",
            "AB": "AK", "AC": "AL", "AD": "AM", "AE": "AO", "AF": "AS", "AG": "AU", "AH": "AX", "AI": "AZ",
            "AJ": "BC", "AK": "BE", "AL": "BG"
        }

        entry_row = target_row
        calc_row = entry_row + 3

        for entry_col, calc_col in entries_to_calculation.items():
            formula = f"=Entries!{entry_col}{entry_row}"
            calc_ws[f"{calc_col}{calc_row}"] = formula

        # Tooling Lookup Formulas
        calc_ws[f"AQ{calc_row}"] = f"=IFNA(VLOOKUP(A{calc_row},Tooling!A:AO,30,0),\"\")"
        calc_ws[f"AR{calc_row}"] = f"=IFNA(VLOOKUP(A{calc_row},Tooling!A:AO,31,0),\"\")"

        wb.save(EXCEL_TEMPLATE)

        # Open the file
        if platform.system() == "Windows":
            os.startfile(EXCEL_TEMPLATE)
        elif platform.system() == "Darwin":
            os.system(f'open "{EXCEL_TEMPLATE}"')
        else:
            os.system(f'xdg-open "{EXCEL_TEMPLATE}"')

        messagebox.showinfo("Success", f"Data {'updated' if overwrite_row else 'saved'} with Sl.# {sl_no} in 'Entries'.")

    except Exception as e:
        messagebox.showerror("Error", str(e))

def run_main_form(root=None, prefill_data=None):
    def update_injection_pressure(event=None):
        raw_material = inputs["Rawmaterial"].get().strip()
        pressure = injection_pressure_map.get(raw_material)

        inputs["Injection Pressure"].config(state="normal")
        inputs["Injection Pressure"].delete(0, tk.END)

        if pressure is not None:
            inputs["Injection Pressure"].insert(0, str(pressure))
        else:
            inputs["Injection Pressure"].insert(0, "Not Found")

        inputs["Injection Pressure"].config(state="readonly")
        calculate_projected_area()

    def calculate_projected_area(event=None):
        try:
            L = float(inputs["L(mm)"].get())
            W = float(inputs["W(mm)"].get())
            proj_area_pct_str = inputs["Projected Area %"].get().replace("%", "")
            proj_pct = float(proj_area_pct_str) / 100.0
            cav_x = int(inputs["No. of cavities -X direction"].get())
            cav_y = int(inputs["No. of cavities -Y direction"].get())
            total = (L * W * 1e-2) * proj_pct * (cav_x * cav_y)
            total = round(total, 2)
            inputs["Total Projected Area"].config(state="normal")
            inputs["Total Projected Area"].delete(0, tk.END)
            inputs["Total Projected Area"].insert(0, str(total))
            inputs["Total Projected Area"].config(state="readonly")
        except:
            inputs["Total Projected Area"].config(state="normal")
            inputs["Total Projected Area"].delete(0, tk.END)
            inputs["Total Projected Area"].insert(0, "Error")
            inputs["Total Projected Area"].config(state="readonly")

    def calculate_tonnage_popup():
        try:
            inj = float(inputs["Injection Pressure"].get())
            area = float(inputs["Total Projected Area"].get())
            tonnage = round((inj * area * 1e-3) * 1.1, 2)
            messagebox.showinfo("Tonnage", f"Tonnage = {tonnage} tons")

            # ✅ Enable machine selection after tonnage is calculated
            inputs["M/c selected"].config(state="readonly")

        except Exception as e:
            messagebox.showerror("Error", f"Check inputs. {e}")

    def submit():
        try:
            entry = {}
            for field in fields:
                val = inputs[field].get().strip()
                if field in dropdowns and not val:
                    messagebox.showerror("Missing Input", f"Select a value for '{field}'")
                    return
                if field in numeric_fields and val:
                    val = numeric_fields[field](val)
                entry[field] = val

            save_to_standard_excel(entry)
            root.destroy()

        except Exception as e:
            messagebox.showerror("Error", str(e))

    global inputs
    inputs = {}

    root = tk.Toplevel()
    root.title("Plastic Part Entry Form")
    root.geometry("1400x1000")

    main_frame = tk.Frame(root)
    main_frame.pack(fill="both", expand=True)

    canvas = tk.Canvas(main_frame)
    scrollbar = tk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)

    scrollbar.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)

    scrollable_frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind_all("<MouseWheel>", _on_mousewheel)

    left_fields = [
        "In-house/Purchase", "ZF Part Number", "Part No", "Part Name", "Volume P.A Nos",
        "No.off (per Vehicle)", "Rawmaterial", "L(mm)", "W(mm)", "H(mm)", "T(mm)",
        "No. of cavities -X direction", "No. of cavities -Y direction",
        "Grain requirement", "HRS Type", "No.of drop",
        "X Direction Slider", "Y Direction Slider",
        "No.of Sliders/Cavity", "No.of Lifters/cavity",
        "Injection Pressure", "Total Projected Area"
    ]

    right_fields = [
        "No.of Bosses /cavity", "No.of Ejector pads/ cavity", "Projected Area %",
        "Finish wt.( Kgs)", "Scrap cost/Kg (INR)", "Bop Parts", "Qty (Nos)",
        "Cost /part (INR)", "Process", "M/c selected", "Machine Cycle time (sec)",
        "Labor Type (Machine)", "Operator /machine", "Labor Type (Setup)",
        "No.of operator", "Time/Lot (min)"
    ]

    printed_sections = set()
    row_left = 0
    row_right = 0

    for field in left_fields:
        section = get_section(field)
        if section and section not in printed_sections:
            tk.Label(scrollable_frame, text=section, font=("Arial", 10, "bold"), anchor="w").grid(
                row=row_left, column=0, columnspan=2, sticky="w", pady=(10, 0)
            )
            printed_sections.add(section)
            row_left += 1

        label = tk.Label(scrollable_frame, text=field, font=("Arial", 10), anchor="w")
        label.grid(row=row_left, column=0, sticky="w", padx=5, pady=2)

        if field in ["Injection Pressure", "Total Projected Area"]:
            inp = tk.Entry(scrollable_frame, width=43, state="readonly")
        elif field in dropdowns:
            inp = ttk.Combobox(scrollable_frame, values=dropdowns[field], width=40, state="readonly")
        else:
            inp = tk.Entry(scrollable_frame, width=43)

        inp.grid(row=row_left, column=1, padx=5, pady=2, sticky="w")
        inputs[field] = inp

        if field == "Rawmaterial":
            inp.bind("<<ComboboxSelected>>", update_injection_pressure)
        if field in ["L(mm)", "W(mm)", "Projected Area %", "No. of cavities -X direction",
                     "No. of cavities -Y direction"]:
            inp.bind("<FocusOut>", calculate_projected_area)

        row_left += 1

    for field in right_fields:
        section = get_section(field)
        if section and section not in printed_sections:
            tk.Label(scrollable_frame, text=section, font=("Arial", 10, "bold"), anchor="w").grid(
                row=row_right, column=3, columnspan=2, sticky="w", pady=(10, 0)
            )
            printed_sections.add(section)
            row_right += 1

        label = tk.Label(scrollable_frame, text=field, font=("Arial", 10), anchor="w")
        label.grid(row=row_right, column=3, sticky="w", padx=5, pady=2)

        if field in ["Injection Pressure", "Total Projected Area"]:
            inp = tk.Entry(scrollable_frame, width=43, state="readonly")
        elif field in dropdowns:
            inp = ttk.Combobox(scrollable_frame, values=dropdowns[field], width=40, state="readonly")
            if field == "M/c selected":
                inp.config(state="disabled")  # Freeze machine selection
        else:
            inp = tk.Entry(scrollable_frame, width=43)

        inp.grid(row=row_right, column=4, padx=5, pady=2, sticky="w")
        inputs[field] = inp

        # Pre-fill and disable if applicable
        if prefill_data and field in prefill_data:
            inp.insert(0, prefill_data[field])
            inp.config(state="disabled")

        if field == "Rawmaterial":
            inp.bind("<<ComboboxSelected>>", update_injection_pressure)
        if field in ["L(mm)", "W(mm)", "Projected Area %", "No. of cavities -X direction",
                     "No. of cavities -Y direction"]:
            inp.bind("<FocusOut>", calculate_projected_area)

        row_right += 1

    tonnage_btn = tk.Button(scrollable_frame, text="Calculate Tonnage", command=calculate_tonnage_popup,
                            bg="blue", fg="white", font=("Arial", 12, "bold"))
    tonnage_btn.grid(row=max(row_left, row_right) + 1, column=1, pady=10)

    submit_btn = tk.Button(scrollable_frame, text="Submit", command=submit,
                           bg="green", fg="white", font=("Arial", 12, "bold"))
    submit_btn.grid(row=max(row_left, row_right) + 2, columnspan=4, pady=20)

    root.mainloop()

# ✅ Check if part exists in Entries sheet
def part_already_exists(zf_part):
    try:
        if not os.path.exists(EXCEL_TEMPLATE):
            return False
        wb = load_workbook(EXCEL_TEMPLATE)
        if "Entries" not in wb.sheetnames:
            return False
        ws = wb["Entries"]
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if row and len(row) > 2:
                if str(row[2]).strip() == zf_part:
                    return True
        return False
    except Exception as e:
        print("Error checking part existence:", e)
        return False

def show_export_options():
    popup = tk.Toplevel()
    popup.title("Export Options")
    popup.geometry("300x150")
    popup.resizable(False, False)
    popup.grab_set()  # Makes the popup modal

    label = tk.Label(popup, text="Choose Export Type", font=("Arial", 12))
    label.pack(pady=10)

    def export_with():
        popup.destroy()
        ask_and_export("with")

    def export_without():
        popup.destroy()
        ask_and_export("without")

    btn1 = tk.Button(popup, text="Export with Formulas", command=export_with, width=25, bg="white", fg="black")
    btn2 = tk.Button(popup, text="Export without Formulas", command=export_without, width=25, bg="white", fg="black")

    btn1.pack(pady=5)
    btn2.pack(pady=5)

def ask_and_export(mode):
    zf_part = simpledialog.askstring("Export", "Enter ZF Part Number to export:")
    if not zf_part:
        return
    if mode == "with":
        export_with_formulas(zf_part)
    elif mode == "without":
        export_without_formulas(zf_part)

# ✅ Small window for basic inputs
def open_initial_entry_window():
    small_win = tk.Toplevel()
    small_win.title("New Entry - Basic Info")
    small_win.geometry("300x150")

    def proceed():
        inhouse = inhouse_var.get()
        zf_part = zf_var.get().strip()

        if not inhouse or not zf_part:
            messagebox.showwarning("Missing", "Please fill all fields")
            return

        if part_already_exists(zf_part):
            messagebox.showerror("Exists", "This ZF Part Number already exists!")
        else:
            small_win.destroy()
            run_main_form(prefill_data={
                "In-house/Purchase": inhouse,
                "ZF Part Number": zf_part
            })

    tk.Label(small_win, text="In-house/Purchase").grid(row=0, column=0)
    inhouse_var = ttk.Combobox(small_win, values=["In-House", "Purchase"], state="readonly")
    inhouse_var.grid(row=0, column=1)

    tk.Label(small_win, text="ZF Part Number").grid(row=1, column=0)
    zf_var = tk.Entry(small_win)
    zf_var.grid(row=1, column=1)

    tk.Button(small_win, text="Next", command=proceed).grid(row=2, columnspan=2, pady=10)

def open_excel_file():
    try:
        if os.path.exists(EXCEL_TEMPLATE):
            if platform.system() == "Windows":
                os.startfile(EXCEL_TEMPLATE)
            elif platform.system() == "Darwin":
                os.system(f'open "{EXCEL_TEMPLATE}"')
            else:
                os.system(f'xdg-open "{EXCEL_TEMPLATE}"')
        else:
            messagebox.showerror("Not Found", "Excel file not found.")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# ✅ Entry type selection window (as Toplevel)
def choose_entry_type(parent_window):
    parent_window.destroy()  # ✅ Close the old root window first

    entry_win = tk.Toplevel()
    entry_win.title("Choose Entry Type")
    entry_win.geometry("300x150")

    tk.Button(entry_win, text="New Entry", width=20,
              command=lambda: [entry_win.destroy(), open_initial_entry_window()]).pack(pady=10)

    tk.Button(entry_win, text="Edit Existing Entry", width=20,
              command=lambda: [entry_win.destroy(), open_excel_file()]).pack(pady=10)

# ✅ Initial mode selection window — only one root window
def select_mode():
    root = tk.Tk()
    root.withdraw()  # ✅ Hide the blank root window

    mode_win = tk.Toplevel()
    mode_win.title("Select Mode")
    mode_win.geometry("300x150")

    tk.Label(mode_win, text="Choose Mode").pack(pady=10)

    tk.Button(mode_win, text="Calculate", width=20,
              command=lambda: choose_entry_type(mode_win)).pack(pady=5)

    tk.Button(mode_win, text="Export", width=20,
              command=lambda: show_export_options()).pack(pady=5)

    root.mainloop()

if __name__ == "__main__":
    select_mode()
